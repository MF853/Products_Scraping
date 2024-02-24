from seleniumbase import Driver
from selenium import webdriver
import random
from selenium.webdriver.common.by import By
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.options import Options
from time import sleep
from selenium.webdriver.support.select import Select
import openpyxl
import random
import pandas as pd
import re

website = 'https://amazon.com/'
proxies = [
    "204.217.194.251:3128:xyzh4525:jpatiosmxoksej",
    "204.217.194.250:3128:xyzh4525:jpatiosmxoksej",
    "204.217.194.249:3128:xyzh4525:jpatiosmxoksej",
    "204.217.194.248:3128:xyzh4525:jpatiosmxoksej",
    "204.217.194.247:3128:xyzh4525:jpatiosmxoksej"
]

# Escolhe um proxy aleatório
proxy = random.choice(proxies)
user_agents = [
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36 Edg/91.0.864.59",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:89.0) Gecko/20100101 Firefox/89.0",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36 OPR/76.0.4017.177",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36 EdgA/91.0.864.59"
]

# Escolhe um agente de usuário aleatório
user_agent = random.choice(user_agents)
arguments = [f'--user-agent={user_agent}']
chrome_options = Options()
chrome_options.add_argument('--headless')
chrome_options.add_argument('--incognito')
chrome_options.add_argument('--disable-gpu')
chrome_options.add_argument('--no-sandbox')
chrome_options.add_argument(arguments[0])

proxy = "xyzh4525:jpatiosmxoksej@204.217.194.247:3128"

driver = webdriver.Chrome(options=chrome_options)
actions = ActionChains(driver)

def natural1(searchtyped, search):
    for letter in searchtyped:
        search.send_keys(letter)
        sleep(random.randint(1, 5) / 30)

site1 = "https://www.tendaatacado.com.br"

# Caminho do arquivo Excel
caminho_arquivo = './uploads/CompraDoMes.xlsx'

# Carregar o arquivo Excel
workbook = openpyxl.load_workbook(caminho_arquivo)

# Selecionar a planilha desejada (pode ser necessário ajustar o nome da planilha)
sheet = workbook.active

# Nomes dos produtos serão armazenados nesta lista
nomes_produtos = []
new_quant = []

# Iterar sobre as células das três primeiras colunas e concatenar os dados
for row in sheet.iter_rows(min_row=2, max_col=3, values_only=True):
    nome_produto = ' '.join(map(str, row))  # Concatena os dados das três primeiras colunas
    nomes_produtos.append(nome_produto)

# Fechar o arquivo Excel após a leitura
workbook.close()

# Adicionar colunas para cada site no arquivo Excel
sheet['D1'] = 'Tenda Atacado'
sheet['E1'] = 'Atacadão'
sheet['F1'] = 'Pão de Açúcar'
sheet['G1'] = 'Mercado Livre'
sheet['H1'] = 'Magazine Luíza'
sheet['I1'] = 'Extra'
sheet['J1'] = 'Carrefour'
sheet['K1'] = 'Sonda Delivery'
sheet['L1'] = 'Ifood'

# Imprimir os nomes dos produtos ou realizar outras operações com eles
novo_cep_aleatorio = "07179-230"

def padronizar_preco(preco_str):
    # Extrair apenas os valores numéricos
    preco_numerico = re.findall(r'\d+,\d+|\d+', preco_str)
    if preco_numerico:
        # Formatar o preço no padrão R$ X,XX
        return f'R$ {preco_numerico[0]}'
    else:
        return ''

padrao = re.compile(r'^(.*?) (\d+(\.\d+)?)\s?(\w+)$')
sites = []
quantidade = []
unidade = []
precos_tenda_atacado = []
precos_atacadao = []
precos_pao_de_acucar = []
precos_mercado_livre = []
precos_magazine_luiza = []
precos_extra = []
precos_carrefour = []
precos_sonda_delivery = []
precos_ifood = []
novo_nome_produto = []

#Tenda atacado
for nome_produto in nomes_produtos:
    novo_nome_produto.append(nome_produto)
    driver.get(f"https://www.tendaatacado.com.br/busca?q=" + nome_produto)
    sites.append("Tenda Atacado")
    driver.maximize_window()
    driver.implicitly_wait(10)
    try:
        zip_code = driver.find_element(By.XPATH, "//input[@name='zipCode']")
        natural1(novo_cep_aleatorio, zip_code)
        exit_buttons = driver.find_elements(By.XPATH, "//img[@class='svgIcon svg-ico_close_with_circle']")
        exit_button = exit_buttons[0]
        exit_button.click()
        driver.implicitly_wait(10)
    except Exception:
        pass
    try:

        produtos_tenda_atacado = driver.find_elements(By.XPATH, "//h3[@class='card-text']")
        info_produto = produtos_tenda_atacado[0].text
        correspondencia = padrao.match(info_produto)
        nome = correspondencia.group(1)
        quantidades = correspondencia.group(2)
        unidades = correspondencia.group(4)
        new_quant.append(quantidades + unidades)
        unidade.append(unidades)
        print(info_produto)
        preço_produtos = driver.find_elements(By.XPATH, "//div[@class='price-block']")
        preço_produto = preço_produtos[0].text
        precos_tenda_atacado.append(padronizar_preco(preço_produto))

        print(preço_produto)
        driver.implicitly_wait(10)
    except Exception:
        if not produtos_tenda_atacado:
            precos_tenda_atacado.append("N/A")
            quantidade.append("N/A")
            unidade.append("N/A")
    # atacadao
    novo_nome_produto.append(nome_produto)
    site = f"https://www.atacadao.com.br/s?q=" + nome_produto + "&sort=score_desc&page=0"
    sites.append("Atacadão")
    driver.get(site)
    driver.maximize_window()
    driver.implicitly_wait(10)
    try:
        clicar_exit = driver.find_element(By.XPATH, "//button[@class='w-5 h-5']")
        clicar_exit.click()
        driver.implicitly_wait(10)
        
    except Exception:
        pass
    try:
        produtos_atacadao = driver.find_elements(By.XPATH, '//a[@data-testid="product-link"]')
        produto_atacadao = produtos_atacadao[0].text
        correspondencia = padrao.match(produto_atacadao)
        nome = correspondencia.group(1)
        print(produto_atacadao)
        preços_produtos_atacadao = driver.find_elements(By.XPATH, "//p[@class='text-2xl text-neutral-500 font-bold']")
        preço_produto_atacadao = preços_produtos_atacadao[0].text
        precos_atacadao.append(padronizar_preco(preço_produto_atacadao))
        print(preço_produto_atacadao)
        driver.implicitly_wait(10)
    except Exception:
        if not produtos_atacadao:
            quantidade.append("N/A")
            unidade.append("N/A")
    # pao de açucar
    sites.append("Pão de Açúcar")
    novo_nome_produto.append(nome_produto)
    site = "https://www.paodeacucar.com/"
    driver.get(site)
    driver.maximize_window()
    driver.implicitly_wait(10)
    search_paodeacucar = driver.find_element(By.XPATH, "//input[@id='input-search']")
    natural1(nome_produto, search_paodeacucar)
    driver.implicitly_wait(10)
    butao_search_pda = driver.find_element(By.XPATH, "//button[@class='search-barstyles__Submit-sc-125xygl-5 kSjsQX buttonstyles__ButtonStyle-sc-1mux0mx-0 dCjjxN']")
    butao_search_pda.click()
    try:
        sleep(6)
        nome_produto_pdas = driver.find_elements(By.XPATH, "//a[@class='product-cardstyles__Link-sc-1uwpde0-9 bSQmwP hyperlinkstyles__Link-j02w35-0 coaZwR']")
        nome_produto_pda = nome_produto_pdas[0].text
        correspondencia = padrao.match(nome_produto_pda)
        nome = correspondencia.group(1)
        print(nome_produto_pda)
        preco_produto_pdas = driver.find_elements(By.XPATH, "//p[@class='price-tag-normalstyle__LabelPrice-sc-1co9fex-0 lkWvql']")
        preco_produto_pda = preco_produto_pdas[0].text
        precos_pao_de_acucar.append(padronizar_preco(preco_produto_pda))
        print(preco_produto_pda)
        driver.implicitly_wait(10)
    except Exception:
        if not nome_produto_pdas:
            quantidade.append("N/A")
            unidade.append("N/A")

    # mercado livre
    novo_nome_produto.append(nome_produto)
    sites.append("Mercado Livre")
    site = "https://www.mercadolivre.com.br/"
    driver.get(site)
    driver.maximize_window()
    driver.implicitly_wait(10)
    search_mercadolivre = driver.find_element(By.XPATH, "//input[@class='nav-search-input']")
    natural1(nome_produto, search_mercadolivre)
    sleep(1)
    butao_search_mercadolivre = driver.find_element(By.XPATH, "//div[@class='nav-icon-search']")
    butao_search_mercadolivre.click()
    driver.implicitly_wait(10)
    try:
        nome_produto_mercadolivres = driver.find_elements(By.XPATH, "//a[@class='ui-search-item__group__element ui-search-link__title-card ui-search-link']")
        nome_produto_mercadolivre = nome_produto_mercadolivres[0].text
        correspondencia = padrao.match(nome_produto_mercadolivre)
        nome = correspondencia.group(1)
        print(nome_produto_mercadolivre)

        preco_produto_mercadolivres = driver.find_elements(By.XPATH, "//span[@class='andes-money-amount ui-search-price__part ui-search-price__part--medium andes-money-amount--cents-superscript']")
        preco_produto_mercadolivre = preco_produto_mercadolivres[0].text
        precos_mercado_livre.append(padronizar_preco(preco_produto_mercadolivre))
        print(preco_produto_mercadolivre)
        driver.implicitly_wait(10)

    except Exception:
        if not nome_produto_mercadolivres:
            precos_mercado_livre.append("N/A")


    # magazine luiza
    sites.append("Magazine Luíza")
    novo_produto = '+'.join(nome_produto.split())
    site = f"https://www.magazineluiza.com.br/busca/{novo_produto}/"
    novo_nome_produto.append(nome_produto)
    driver.get(site)
    sleep(3)
    driver.maximize_window()
    driver.implicitly_wait(10)
    try:
        nome_produto_mercadolivres = driver.find_elements(By.XPATH, '//h2[@data-testid="product-title"]')
        nome_produto_mercadolivre = nome_produto_mercadolivres[0].text
        correspondencia = padrao.match(nome_produto_mercadolivre)
        nome = correspondencia.group(1)
        print(nome_produto_mercadolivre)
        preco_produto_mercadolivres = driver.find_elements(By.XPATH, '//p[@data-testid="price-value"]')
        preco_produto_mercadolivre = preco_produto_mercadolivres[0].text
        precos_magazine_luiza.append(padronizar_preco(preco_produto_mercadolivre))
        print(preco_produto_mercadolivre)
        driver.implicitly_wait(10)
    except Exception:
        if not nome_produto_mercadolivres:
            quantidade.append("N/A")
            unidade.append("N/A")

    # extra
    sites.append("Extra")
    novo_produto_atacadao = '%20'.join(nome_produto.split())
    site = f"https://www.clubeextra.com.br/busca?terms={novo_produto_atacadao}" 
    novo_nome_produto.append(nome_produto)
    driver.get(site)
    driver.maximize_window()
    driver.implicitly_wait(10)
    try:
        nome_produto_mercadolivres = driver.find_elements(By.XPATH, "//a[@class='product-cardstyles__Link-sc-1uwpde0-9 bSQmwP hyperlinkstyles__Link-j02w35-0 iMacJg']")
        nome_produto_mercadolivre = nome_produto_mercadolivres[0].text
        correspondencia = padrao.match(nome_produto_mercadolivre)
        nome = correspondencia.group(1)
        print(nome_produto_mercadolivre)
        preco_produto_mercadolivres = driver.find_elements(By.XPATH, "//p[@class='price-tag-normalstyle__LabelPrice-sc-1co9fex-0 lkWvql']")
        preco_produto_mercadolivre = preco_produto_mercadolivres[0].text
        precos_extra.append(padronizar_preco(preco_produto_mercadolivre))
        print(preco_produto_mercadolivre)
        driver.implicitly_wait(10)
    except Exception:
        if not nome_produto_mercadolivres:
            quantidade.append("N/A")
            unidade.append("N/A")


    # carrefour
    sites.append("Carrefour")
    novo_produto_carrefour = '+'.join(nome_produto.split())
    site = f"https://mercado.carrefour.com.br/s?q={novo_produto_carrefour}&sort=score_desc&page=0"
    novo_nome_produto.append(nome_produto)
    driver.get(site)
    driver.maximize_window()
    driver.implicitly_wait(10)
    try:
        choose = driver.find_elements(By.XPATH, "//div[@class='border border-gray-300 items-center flex flex-col text-center rounded-lg shadow-md hover:ring-1 ring-slate-300 pointer-events-auto select-none cursor-pointer p-1 w-1/2 sm:w-1/3 h-40 pt-[21px] max-w-[128px] overflow-hidden']")
        choose_correct = choose[1]
        choose_correct.click()
        choose_city_dropdown = driver.find_element(By.XPATH, "//select[@id='selectCity']")
        choose_city_dropdown.click()
        choose_maceio = driver.find_element(By.XPATH, '//option[@value="Maceió"]')
        choose_maceio.click()
        choose_mercado = driver.find_element(By.XPATH, "//div[@class='grid grid-flow-col']")
        choose_mercado.click()
    except Exception:
        pass
    try:
        nome_produto_mercadolivres = driver.find_elements(By.XPATH, "//span[@class='overflow-hidden text-ellipsis -webkit-box -webkit-line-clamp-3 -webkit-box-orient-vertical text-[13px] text-monet-400 ']")
        nome_produto_mercadolivre = nome_produto_mercadolivres[0].text
        correspondencia = padrao.match(nome_produto_mercadolivre)
        nome = correspondencia.group(1)
        print(nome_produto_mercadolivre)
        preco_produto_mercadolivres = driver.find_elements(By.XPATH, '//span[@data-test-id="price"]')
        preco_produto_mercadolivre = preco_produto_mercadolivres[0].text
        precos_carrefour.append(padronizar_preco(preco_produto_mercadolivre))
        print(preco_produto_mercadolivre)
        driver.implicitly_wait(10)
    except Exception:
        if not nome_produto_mercadolivres:
            quantidade.append("N/A")
            unidade.append("N/A")

    #sonda delivery
    sites.append("Sonda Delivery")
    novo_produto_carrefour = '%20'.join(nome_produto.split())
    site = f"https://sondadelivery.com.br/delivery/busca/{novo_produto_carrefour}"
    novo_nome_produto.append(nome_produto)
    driver.get(site)
    driver.maximize_window()
    driver.implicitly_wait(10)
    try:
        nome_produto_mercadolivres = driver.find_elements(By.XPATH, "//h3[@class='product--title']")
        nome_produto_mercadolivre = nome_produto_mercadolivres[0].text
        correspondencia = padrao.match(nome_produto_mercadolivre)
        nome = correspondencia.group(1)

        print(nome_produto_mercadolivre)
        preco_produto_mercadolivres = driver.find_elements(By.XPATH, "//span[@class='price']")
        preco_produto_mercadolivre = preco_produto_mercadolivres[0].text
        precos_sonda_delivery.append(padronizar_preco(preco_produto_mercadolivre))
        print(preco_produto_mercadolivre)
        driver.implicitly_wait(10)
    except Exception:
        if not nome_produto_mercadolivres:
            quantidade.append("N/A")
            unidade.append("N/A")


    # ifood
    sites.append("Ifood")
    novo_produto_carrefour = '%20'.join(nome_produto.split())
    adress = "rua salma tacache pirilo"
    site = f"https://www.ifood.com.br/busca?q={novo_produto_carrefour}&tab=1"
    novo_nome_produto.append(nome_produto)
    driver.get(site)
    sleep(2)
    driver.maximize_window()
    driver.implicitly_wait(10)
    try:
        search_adresses = driver.find_elements(By.XPATH, "https://www.ifood.com.br/inicio")
        search_adress = search_adresses[1]
        natural1(adress, search_adress)
        driver.implicitly_wait(10)
        click_adresses = driver.find_elements(By.XPATH, "//button[@class='btn-address--full-size']")
        click_adress= click_adresses[1]
        click_adress.click()
        driver.implicitly_wait(10)
        choose_local = driver.find_element(By.XPATH, "//button[@class='btn btn--default btn--size-m address-maps__submit']")
        choose_local.click()
        driver.implicitly_wait(10)
        salvar = driver.find_element(By.XPATH, "//button[@type='submit']")
        salvar.click()
        driver.implicitly_wait(10)
    except Exception:
        pass
    try:
        nome_produto_mercadolivres = driver.find_elements(By.XPATH, "//h4[@class='merchant-list-carousel__item-title']")
        nome_produto_mercadolivre = nome_produto_mercadolivres[0].text
        correspondencia = padrao.match(nome_produto_mercadolivre)
        nome = correspondencia.group(1)
        print(f"Produto IFOOD:{nome_produto_mercadolivre}")
        driver.implicitly_wait(10)
        preco_produto_mercadolivres = driver.find_elements(By.XPATH, "//span[@class='card-stack-item-price--regular card-stack-price__DEFAULT']")
        preco_produto_mercadolivre = preco_produto_mercadolivres[0].text
        precos_ifood.append(padronizar_preco(preco_produto_mercadolivre))
        print(f"Preço IFOOD{preco_produto_mercadolivre}")
        driver.implicitly_wait(10)
    except Exception:
        if not nome_produto_mercadolivres:
            quantidade.append("N/A")
            unidade.append("N/A")

    # Garantir que todas as listas tenham o mesmo comprimento
    # Estender as listas para garantir que todas tenham o mesmo comprimento
    sheet = workbook.active

# Garantir que todas as listas tenham o mesmo comprimento antes de começar a salvar os dados
# Isso pode ser necessário se houver alguma inconsistência no número de produtos encontrados por site
    max_length = max(len(precos_tenda_atacado), len(precos_atacadao), len(precos_pao_de_acucar), len(precos_mercado_livre), len(precos_magazine_luiza), len(precos_extra), len(precos_carrefour), len(precos_sonda_delivery), len(precos_ifood))

    # Preencher as listas para garantir que todas tenham o mesmo comprimento
    lists = [precos_tenda_atacado, precos_atacadao, precos_pao_de_acucar, precos_mercado_livre, precos_magazine_luiza, precos_extra, precos_carrefour, precos_sonda_delivery, precos_ifood]
    for lista in lists:
        while len(lista) < max_length:
            lista.append("N/A") # Adiciona "N/A" para produtos não encontrados ou erros

    # Salvar os dados no Excel
    for i in range(max_length):
        sheet.cell(row=i + 2, column=4).value = precos_tenda_atacado[i]
        sheet.cell(row=i + 2, column=5).value = precos_atacadao[i]
        sheet.cell(row=i + 2, column=6).value = precos_pao_de_acucar[i]
        sheet.cell(row=i + 2, column=7).value = precos_mercado_livre[i]
        sheet.cell(row=i + 2, column=8).value = precos_magazine_luiza[i]
        sheet.cell(row=i + 2, column=9).value = precos_extra[i]
        sheet.cell(row=i + 2, column=10).value = precos_carrefour[i]
        sheet.cell(row=i + 2, column=11).value = precos_sonda_delivery[i]
        sheet.cell(row=i + 2, column=12).value = precos_ifood[i]

    # Salvar o arquivo Excel após sair do loop de iteração de produtos
    workbook.save(caminho_arquivo)